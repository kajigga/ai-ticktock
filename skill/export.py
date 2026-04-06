#!/usr/bin/env python3
# export.py — export time entries to the Excel timesheet (Log-based v5 format)
#
# Usage:
#   Normal mode: uv run --with openpyxl python3 export.py [WEEK_START]
#     WEEK_START = Monday date in YYYY-MM-DD (defaults to current week).
#     Replaces all Log entries for that week, then ensures Summary has a
#     column for the week with SUMIFS formulas.
#
#   All mode:   uv run --with openpyxl python3 export.py --all
#     Clears the entire Log sheet and writes every entry from all time.
#     Also refreshes the Summary columns for every week present.
#
#   Setup mode: uv run --with openpyxl python3 export.py --setup
#     One-time migration: consolidates all existing week tabs → Log,
#     updates Settings with the current account list, and rewrites
#     Summary so it pulls from Log via SUMIFS.
#
# Reads ~/.config/timetracker.json for data_file and spreadsheet_file.

import json, os, re, shutil, sys, tempfile
from datetime import datetime, timedelta
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def safe_save(wb, dest_path: str) -> None:
    """Save workbook via a temp file then atomic copy to avoid OneDrive sync races."""
    fd, tmp = tempfile.mkstemp(suffix='.xlsx', dir='/tmp')
    os.close(fd)
    wb.save(tmp)
    shutil.copy2(tmp, dest_path)
    os.unlink(tmp)

# ── Account list (must match timetracker accounts) ───────────────────────────
ACCOUNTS = [
    'GXO', 'FCB', 'Liberty Mutual', 'Annual Leave',
    'Internal - Non Billable', 'Acrisure', 'Acrisure P1',
    'Arrow', 'Barracuda', 'IHG', 'Snowflake', 'QTS',
]

# ── Activity type keyword map (checked against notes, highest match wins) ────
ACTIVITY_KEYWORDS = [
    (['workshop', 'working session', 'standup', 'stand-up', 'sync', 'meeting', 'call', 'review'], 'Workshop/Working Session'),
    (['develop', 'config', 'build', 'implement', 'integrat', 'code', 'script', 'deploy'], 'Development/Configuration'),
    (['analy', 'query', 'report', 'investigat', 'research'], 'Analysis/Query Building'),
    (['test', 'uat', 'qa', 'debug', 'validat'], 'Testing/UAT'),
    (['project manag', 'kickoff', 'kick-off', 'planning', 'status', 'deck', 'slide'], 'Project Management'),
    (['doc', 'write up', 'notes', 'runbook'], 'Documentation'),
    (['train', 'enabl', 'onboard', 'demo', 'workshop'], 'Training/Enablement'),
    (['travel'], 'Travel'),
]

# ── Style helpers ────────────────────────────────────────────────────────────
def header_style(cell, bg='1F3864', fg='FFFFFF'):
    """Apply dark-header styling (bold white text on dark bg, centered)."""
    cell.font = Font(bold=True, color=fg, name='Calibri', size=10)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center')

# ── Helpers ───────────────────────────────────────────────────────────────────

def infer_activity_type(work_type: str, notes: str) -> str:
    """Map timetracker work_type + notes text to an Excel activity type label."""
    if work_type in ('Non-Billable', 'PTO'):
        return 'Internal/Admin'
    notes_lower = (notes or '').lower()
    for keywords, activity in ACTIVITY_KEYWORDS:
        if any(kw in notes_lower for kw in keywords):
            return activity
    return 'Workshop/Working Session'


def parse_tab_week_start(tab_name: str, title_value) -> Optional[datetime]:
    """
    Derive the Monday week-start date for an old week tab.
    Tries the title cell first (e.g. 'Week of Mar 31, 2026'), then the
    tab name ('31Mar2026' or 'Wk Apr 14').  Returns None on failure.
    """
    # Attempt 1: parse 'Week of <Month> <Day>, <Year>' from title cell
    if title_value:
        text = str(title_value)
        m = re.search(r'(\w+\s+\d+,?\s*\d{4})', text)
        if m:
            date_str = m.group(1).replace(',', '')
            for fmt in ('%b %d %Y', '%B %d %Y'):
                try:
                    return datetime.strptime(date_str, fmt)
                except ValueError:
                    pass

    # Attempt 2: tab name like '31Mar2026'
    try:
        return datetime.strptime(tab_name, '%d%b%Y')
    except ValueError:
        pass

    # Attempt 3: tab name like 'Wk Apr 14' (assume year 2026)
    m = re.match(r'Wk\s+(\w+)\s+(\d+)$', tab_name.strip())
    if m:
        try:
            dt = datetime.strptime(f"{m.group(1)} {m.group(2)} 2026", '%b %d %Y')
            # Snap to Monday of that week
            return dt - timedelta(days=dt.weekday())
        except ValueError:
            pass

    return None


def find_or_create_log(wb) -> openpyxl.worksheet.worksheet.Worksheet:
    """Return the Log sheet, creating it (with headers) if absent."""
    if 'Log' in wb.sheetnames:
        return wb['Log']
    # Insert after Settings if it exists, otherwise at position 2
    insert_pos = 2
    if 'Settings' in wb.sheetnames:
        insert_pos = wb.sheetnames.index('Settings') + 1
    log_ws = wb.create_sheet('Log', insert_pos)
    for col_letter, width in zip('ABCDEFG', [34, 13, 13, 24, 28, 7, 42]):
        log_ws.column_dimensions[col_letter].width = width
    for col, header in enumerate(['Entry ID', 'Date', 'Week Start', 'Customer', 'Activity Type', 'Hours', 'Notes'], 1):
        header_style(log_ws.cell(1, col, header))
    return log_ws


def ensure_summary_column(wb, week_start_dt: datetime) -> None:
    """
    Ensure the Summary sheet has a column for week_start_dt in row 3,
    then write SUMIFS formulas for every customer row (col A non-empty).
    """
    if 'Summary' not in wb.sheetnames:
        return
    sum_ws = wb['Summary']

    # Scan row 3 for an existing column matching this week start
    week_col = None
    last_used_col = 1
    max_scan = max(sum_ws.max_column + 1, 20)
    for col in range(2, max_scan + 1):
        val = sum_ws.cell(3, col).value
        if val is None:
            if week_col is None:
                week_col = col   # first empty slot becomes the new column
            break
        last_used_col = col
        if isinstance(val, datetime) and val.date() == week_start_dt.date():
            week_col = col
            break
        # Also try string comparison
        try:
            if datetime.strptime(str(val), '%Y-%m-%d').date() == week_start_dt.date():
                week_col = col
                break
        except (ValueError, TypeError):
            pass

    if week_col is None:
        week_col = last_used_col + 1

    # Write / refresh the week header in row 3
    hdr = sum_ws.cell(3, week_col, week_start_dt)
    hdr.number_format = 'DD MMM YYYY'
    hdr.font = Font(bold=True, name='Calibri', size=10)
    hdr.alignment = Alignment(horizontal='center')

    # Write SUMIFS formulas for each non-empty customer row
    col_letter = get_column_letter(week_col)
    for row_idx in range(4, 30):
        customer = sum_ws.cell(row_idx, 1).value
        if not customer:
            continue
        formula = (
            f'=IF($A{row_idx}="",'
            f'"",SUMIFS(Log!$F:$F,Log!$D:$D,$A{row_idx},Log!$C:$C,{col_letter}$3))'
        )
        sum_ws.cell(row_idx, week_col).value = formula


# ── Load config ───────────────────────────────────────────────────────────────
config_path = os.path.expanduser('~/.config/timetracker.json')
with open(config_path) as f:
    config = json.load(f)

data_file = config['data_file']
xlsx_path = os.path.expanduser(config['spreadsheet_file'])

if not os.path.exists(xlsx_path):
    print(f"Error: workbook not found at {xlsx_path}", file=sys.stderr)
    sys.exit(1)

# ── Determine mode ────────────────────────────────────────────────────────────
setup_mode = '--setup' in sys.argv
all_mode   = '--all'   in sys.argv
if setup_mode or all_mode:
    week_start_dt = None
    week_start_str = None
else:
    if len(sys.argv) >= 2:
        week_start_str = sys.argv[1]
    else:
        today = datetime.today()
        week_start_str = (today - timedelta(days=today.weekday())).strftime('%Y-%m-%d')
    week_start_dt = datetime.strptime(week_start_str, '%Y-%m-%d')

wb = openpyxl.load_workbook(xlsx_path)

# ════════════════════════════════════════════════════════════════════════════
# SETUP MODE — one-time migration
# ════════════════════════════════════════════════════════════════════════════
if setup_mode:
    SKIP = {'Summary', 'Settings', 'Log'}
    week_tabs = [s for s in wb.sheetnames if s not in SKIP]

    # 1. Build (or recreate) the Log sheet ------------------------------------
    if 'Log' in wb.sheetnames:
        del wb['Log']
    log_ws = find_or_create_log(wb)

    # 2. Migrate data from each week tab → Log --------------------------------
    migrated = 0
    skipped_tabs = []
    log_rows = []   # (date_dt, ws_dt, customer, activity, hours, notes)

    for tab_name in week_tabs:
        ws = wb[tab_name]
        ws_dt = parse_tab_week_start(tab_name, ws.cell(1, 1).value)
        if ws_dt is None:
            skipped_tabs.append(tab_name)
            continue
        # Data starts at row 3 (row 1 = title, row 2 = column headers)
        for row in ws.iter_rows(min_row=3, max_col=4, values_only=True):
            customer, activity, hours, notes = row
            if not customer or not hours:
                continue
            try:
                h = float(hours)
            except (TypeError, ValueError):
                continue
            log_rows.append((
                ws_dt,
                ws_dt,
                str(customer).strip(),
                str(activity).strip() if activity else 'Workshop/Working Session',
                h,
                str(notes).strip() if notes else '',
            ))
            migrated += 1

    # Sort by week start, then customer
    log_rows.sort(key=lambda r: (r[1], r[2]))
    for i, (date_dt, ws_dt, customer, activity, hours, notes) in enumerate(log_rows, 2):
        log_ws.cell(i, 1, date_dt).number_format = 'YYYY-MM-DD'
        log_ws.cell(i, 2, ws_dt).number_format  = 'YYYY-MM-DD'
        log_ws.cell(i, 3, customer)
        log_ws.cell(i, 4, activity)
        log_ws.cell(i, 5, hours)
        log_ws.cell(i, 6, notes)

    print(f"  Log: migrated {migrated} rows from {len(week_tabs) - len(skipped_tabs)} tabs")
    if skipped_tabs:
        print(f"  Warning: could not parse week start for tabs: {', '.join(skipped_tabs)}")

    # 3. Update Settings!A4:A(4+N) with Kevin's accounts ---------------------
    if 'Settings' in wb.sheetnames:
        settings_ws = wb['Settings']
        for i, account in enumerate(ACCOUNTS, start=4):
            settings_ws.cell(i, 1, account)
        # Clear any leftover placeholder rows beyond the list
        for i in range(len(ACCOUNTS) + 4, 25):
            settings_ws.cell(i, 1, None)
        print(f"  Settings: wrote {len(ACCOUNTS)} accounts to A4:A{len(ACCOUNTS)+3}")
    else:
        print("  Warning: Settings sheet not found — accounts not updated")

    # 4. Rewrite Summary formulas to pull from Log ----------------------------
    if 'Summary' in wb.sheetnames:
        sum_ws = wb['Summary']

        # Collect unique week-start dates from migrated data, sorted
        unique_weeks = sorted({r[1] for r in log_rows})

        # Write customer names from ACCOUNTS into column A (rows 4+)
        sum_ws.cell(3, 1, 'Customer')
        sum_ws.cell(3, 1).font = Font(bold=True, name='Calibri', size=10)

        for i, account in enumerate(ACCOUNTS, start=4):
            sum_ws.cell(i, 1, account)
        # Clear rows beyond the account list
        for i in range(len(ACCOUNTS) + 4, 25):
            sum_ws.cell(i, 1, None)

        # Write week-start dates across row 3 (cols B, C, D…)
        # Reuse existing columns if their dates match, otherwise overwrite
        for col_offset, ws_dt in enumerate(unique_weeks):
            col = col_offset + 2
            c = sum_ws.cell(3, col, ws_dt)
            c.number_format = 'DD MMM YYYY'
            c.font = Font(bold=True, name='Calibri', size=10)
            c.alignment = Alignment(horizontal='center')
            col_letter = get_column_letter(col)
            for row_idx in range(4, len(ACCOUNTS) + 4):
                formula = (
                    f'=IF($A{row_idx}="",'
                    f'"",SUMIFS(Log!$F:$F,Log!$D:$D,$A{row_idx},Log!$C:$C,{col_letter}$3))'
                )
                sum_ws.cell(row_idx, col).value = formula

        # Clear any extra week columns beyond what we wrote
        for col in range(len(unique_weeks) + 2, sum_ws.max_column + 1):
            sum_ws.cell(3, col, None)
            for row_idx in range(4, 25):
                sum_ws.cell(row_idx, col, None)

        print(f"  Summary: rewrote {len(ACCOUNTS)} customers × {len(unique_weeks)} weeks with SUMIFS formulas")
    else:
        print("  Warning: Summary sheet not found — formulas not updated")

    # 5. Delete old week tabs -------------------------------------------------
    for tab_name in week_tabs:
        if tab_name in wb.sheetnames:
            del wb[tab_name]
    print(f"  Deleted {len(week_tabs)} old week tabs")

    safe_save(wb, xlsx_path)
    print(f"\nSetup complete → {xlsx_path}")
    sys.exit(0)

# ════════════════════════════════════════════════════════════════════════════
# ALL MODE — clear Log and write every entry from all time
# ════════════════════════════════════════════════════════════════════════════
if all_mode:
    with open(data_file) as f:
        all_entries = json.load(f)

    # Recreate Log sheet from scratch (delete + recreate preserves position)
    if 'Log' in wb.sheetnames:
        log_pos = wb.sheetnames.index('Log')
        del wb['Log']
        log_ws = wb.create_sheet('Log', log_pos)
    else:
        log_ws = find_or_create_log(wb)

    # Apply column widths and headers
    for col_letter, width in zip('ABCDEFG', [34, 13, 13, 24, 28, 7, 42]):
        log_ws.column_dimensions[col_letter].width = width
    for col, header in enumerate(['Entry ID', 'Date', 'Week Start', 'Customer', 'Activity Type', 'Hours', 'Notes'], 1):
        header_style(log_ws.cell(1, col, header))

    # Write all entries sorted by date then project
    sorted_entries = sorted(all_entries, key=lambda e: (e.get('date', ''), e.get('project', '')))
    for i, e in enumerate(sorted_entries, 2):
        entry_dt   = datetime.strptime(e['date'], '%Y-%m-%d')
        ws_str     = e.get('week_start', '')
        ws_dt      = datetime.strptime(ws_str, '%Y-%m-%d') if ws_str else entry_dt
        activity   = infer_activity_type(e.get('work_type', 'Billable'), e.get('notes', ''))
        log_ws.cell(i, 1, e.get('id', ''))
        log_ws.cell(i, 2, entry_dt).number_format = 'YYYY-MM-DD'
        log_ws.cell(i, 3, ws_dt).number_format    = 'YYYY-MM-DD'
        log_ws.cell(i, 4, e['project'])
        log_ws.cell(i, 5, activity)
        log_ws.cell(i, 6, e['hours'])
        log_ws.cell(i, 7, e.get('notes', ''))

    # Refresh Summary columns for every week present
    unique_weeks = sorted({datetime.strptime(e['week_start'], '%Y-%m-%d')
                           for e in all_entries if e.get('week_start')})
    for ws_dt in unique_weeks:
        ensure_summary_column(wb, ws_dt)

    safe_save(wb, xlsx_path)
    total_h = sum(e['hours'] for e in all_entries)
    print(f"Exported all {len(all_entries)} entries ({total_h:.1f}h) across "
          f"{len(unique_weeks)} weeks to {xlsx_path}")
    sys.exit(0)

# ════════════════════════════════════════════════════════════════════════════
# NORMAL EXPORT MODE — write a week's entries to the Log sheet
# ════════════════════════════════════════════════════════════════════════════

# Load entries for the target week
with open(data_file) as f:
    all_entries = json.load(f)
week_entries = [e for e in all_entries if e.get('week_start') == week_start_str]

# 1. Get or create Log sheet --------------------------------------------------
log_ws = find_or_create_log(wb)

# 2. Delete existing rows for this week start ---------------------------------
rows_to_delete = []
for row in log_ws.iter_rows(min_row=2):
    cell_c = row[2].value   # column C = Week Start
    if cell_c is None:
        continue
    if isinstance(cell_c, datetime):
        cell_date = cell_c.date()
    else:
        try:
            cell_date = datetime.strptime(str(cell_c), '%Y-%m-%d').date()
        except ValueError:
            continue
    if cell_date == week_start_dt.date():
        rows_to_delete.append(row[0].row)

for row_num in sorted(rows_to_delete, reverse=True):
    log_ws.delete_rows(row_num)

# 3. Append new entries -------------------------------------------------------
for e in sorted(week_entries, key=lambda e: (e['date'], e['project'])):
    entry_dt = datetime.strptime(e['date'], '%Y-%m-%d')
    activity = infer_activity_type(e.get('work_type', 'Billable'), e.get('notes', ''))
    next_row = log_ws.max_row + 1
    log_ws.cell(next_row, 1, e.get('id', ''))
    log_ws.cell(next_row, 2, entry_dt).number_format = 'YYYY-MM-DD'
    log_ws.cell(next_row, 3, week_start_dt).number_format = 'YYYY-MM-DD'
    log_ws.cell(next_row, 4, e['project'])
    log_ws.cell(next_row, 5, activity)
    log_ws.cell(next_row, 6, e['hours'])
    log_ws.cell(next_row, 7, e.get('notes', ''))

# 4. Ensure Summary has a column for this week with SUMIFS formulas -----------
ensure_summary_column(wb, week_start_dt)

# ── Save ──────────────────────────────────────────────────────────────────────
safe_save(wb, xlsx_path)
total_h = sum(e['hours'] for e in week_entries)
print(f"Exported {len(week_entries)} entries ({total_h:.1f}h) for week of {week_start_str} to {xlsx_path}")
